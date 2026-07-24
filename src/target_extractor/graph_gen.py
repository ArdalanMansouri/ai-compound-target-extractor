import math
import re
from collections import Counter

import pandas as pd
from plotly.subplots import make_subplots
import plotly.graph_objects as go


class Graph:
    """Generates a two-panel keyword-distribution chart from a compound DataFrame.

    The top panel shows the total count per keyword; the bottom panel shows a
    100 % stacked bar chart split by compound type (Normal, Inhibitor, Inducer).

    All visual parameters are exposed as instance attributes so they can be
    changed at any time before calling :meth:`plot`.

    Attributes:
        df (pd.DataFrame): Input DataFrame containing compound and keyword data.
        type_col (str): Column used to group rows by compound type.
        keyword_col (str): Column whose values are parsed for keywords.
        height (int): Figure height in pixels.
        width (int): Figure width in pixels.
        margin (dict): Plotly margin dict with keys ``l``, ``r``, ``t``, ``b``.
        bargap (float): Gap between adjacent bars, expressed as a fraction (0–1).
        vertical_spacing (float): Vertical gap between the two subplot rows.
        row_heights (list[float]): Two-element list with relative heights of the
            top and bottom rows.
        colors (dict[str, str]): Mapping from type label to bar colour string.
    """

    def __init__(
        self,
        df: pd.DataFrame,
        *,
        type_col: str = "Screen: Effect on EV uptake",
        keyword_col: str = "Keywords",
        height: int = 900,
        width: int = 1100,
        margin: dict | None = None,
        bargap: float = 0.15,
        vertical_spacing: float = 0.05,
        row_heights: list | None = None,
        colors: dict | None = None,
    ) -> None:
        """Initialises Graph with a DataFrame and optional visual parameters.

        Args:
            df: Input DataFrame containing compound and keyword data.
            type_col: Column used to group compounds by type (Normal,
                Inhibitor, Inducer).
            keyword_col: Column whose values are parsed for keywords.
            height: Figure height in pixels.
            width: Figure width in pixels.
            margin: Plotly margin dict with keys ``l``, ``r``, ``t``, ``b``.
                Defaults to ``{"l": 60, "r": 40, "t": 90, "b": 180}``.
            bargap: Gap between adjacent bars, expressed as a fraction (0–1).
            vertical_spacing: Vertical gap between the two subplot rows.
            row_heights: Two-element list with relative heights of the top and
                bottom rows. Defaults to ``[0.35, 0.65]``.
            colors: Mapping from type label to bar colour string. Defaults to
                ``{"Normal": "black", "Inhibitor": "green", "Inducer": "red"}``.
        """
        self.df = df
        self.type_col = type_col
        self.keyword_col = keyword_col
        self.height = height
        self.width = width
        self.margin = margin or {"l": 60, "r": 40, "t": 90, "b": 180}
        self.bargap = bargap
        self.vertical_spacing = vertical_spacing
        self.row_heights = row_heights or [0.35, 0.65]
        self.colors = colors or {
            "Normal": "black",
            "Inhibitor": "green",
            "Inducer": "red",
        }

    # ------------------------------------------------------------------
    # Helper methods
    # ------------------------------------------------------------------

    def keyword_counts(
        self, df: pd.DataFrame, col: str | None = None
    ) -> tuple[list, list, list]:
        """Counts keyword occurrences in a DataFrame column.

        Multi-keyword strings are split on ``;``, ``,``, ``|``, or newlines.
        Keywords are returned in case-insensitive alphabetical order.

        Args:
            df: DataFrame to analyse.
            col: Column name to parse. Defaults to :attr:`keyword_col`.

        Returns:
            A three-element tuple ``(keys, counts, percentages)`` where

            * ``keys`` – sorted list of unique keyword strings,
            * ``counts`` – absolute occurrence count for each keyword,
            * ``percentages`` – occurrence count as a percentage of ``len(df)``.
        """
        col = col or self.keyword_col
        s = df[col].dropna().astype(str)
        parts = s.apply(lambda x: re.split(r"[;,|\n]", x))
        flat = [p.strip() for sub in parts for p in sub if p.strip()]
        counts = Counter(flat)
        total = len(df)
        keys_sorted = sorted(counts.keys(), key=lambda v: v.lower())
        counts_sorted = [counts[k] for k in keys_sorted]
        perc_sorted = [c / total * 100 for c in counts_sorted]
        return keys_sorted, counts_sorted, perc_sorted

    def keyword_counts_by_type(
        self, df: pd.DataFrame | None = None
    ) -> dict[str, tuple]:
        """Computes keyword counts separately for each compound type.

        Args:
            df: DataFrame to analyse. Defaults to :attr:`df`.

        Returns:
            A dict with keys ``"Normal"``, ``"Inhibitor"``, and ``"Inducer"``,
            each mapping to the ``(keys, counts, percentages)`` tuple returned
            by :meth:`keyword_counts`.
        """
        df = df if df is not None else self.df
        return {
            t: self.keyword_counts(df[df[self.type_col].astype(str).str.lower() == t.lower()])
            for t in ("Normal", "Inhibitor", "Inducer")
        }

    @staticmethod
    def to_map(keys: list, counts: list) -> dict:
        """Converts parallel key and count lists into a dictionary.

        Args:
            keys: List of keyword strings.
            counts: List of corresponding count values.

        Returns:
            A dict mapping each key to its count.
        """
        return dict(zip(keys, counts))

    @staticmethod
    def to_pct(vals: list, totals: list) -> list:
        """Converts raw counts to per-keyword percentages.

        Each element is divided by its corresponding total so that every
        keyword column sums to 100 %.

        Args:
            vals: Raw counts for one compound type.
            totals: Total counts across all types for each keyword.

        Returns:
            List of percentage values (0–100), with 0 where the total is 0.
        """
        return [(v / t * 100) if t else 0 for v, t in zip(vals, totals)]

    # ------------------------------------------------------------------
    # Private static helpers (shared by scored_bar_plot and scored_table)
    # ------------------------------------------------------------------

    @staticmethod
    def _compounds_for_keyword(
        df: pd.DataFrame, keyword_col: str, kw: str
    ) -> pd.DataFrame:
        """Return the subset of *df* whose *keyword_col* cell contains *kw*.

        Multi-keyword cells are split on ``;``, ``,``, ``|``, or newlines
        before checking for membership.

        Args:
            df: Full compound DataFrame.
            keyword_col: Column whose values are parsed for keywords.
            kw: Keyword to search for.

        Returns:
            Filtered :class:`pandas.DataFrame` containing only matching rows.
        """
        mask = df[keyword_col].apply(
            lambda x: False
            if pd.isna(x)
            else kw in [p.strip() for p in re.split(r"[;,|\n]", str(x))]
        )
        return df[mask]

    @staticmethod
    def _compound_data_list(
        df_sub: pd.DataFrame,
        avg_controls: dict,
        uptake_col: str,
        library_col: str,
        name_col: str,
    ) -> list:
        """Return ``(log2FC, compound_name, ev_uptake)`` tuples sorted ascending.

        Rows whose library is not in *avg_controls*, whose uptake value cannot
        be parsed, or whose uptake is non-positive are silently skipped.

        Args:
            df_sub: Slice of the compound DataFrame for one type/keyword.
            avg_controls: Mapping from library name to average control uptake.
            uptake_col: Column name containing EV-uptake values.
            library_col: Column name containing library identifiers.
            name_col: Column name containing compound names.

        Returns:
            List of ``(log2FC, name, ev_uptake)`` tuples sorted by log2FC
            ascending (most negative first).
        """
        results = []
        for _, row in df_sub.iterrows():
            lib = row.get(library_col, None)
            if lib not in avg_controls:
                continue
            try:
                val = float(row[uptake_col])
            except (TypeError, ValueError, KeyError):
                continue
            if val <= 0:
                continue
            fc = math.log2(val / avg_controls[lib])
            cname = row.get(name_col, "Unknown")
            results.append((fc, str(cname), val))
        return sorted(results, key=lambda t: t[0])

    @staticmethod
    def _percentile_val(data: list, p: float) -> float:
        """Return the p-th percentile (0–100) of a sorted or unsorted list."""
        s = sorted(data)
        n = len(s)
        if n == 1:
            return s[0]
        idx = p / 100 * (n - 1)
        lo = int(idx)
        hi = min(lo + 1, n - 1)
        return s[lo] + (idx - lo) * (s[hi] - s[lo])

    @staticmethod
    def _interpolate_color(
        fc: float,
        fc_min: float,
        fc_max: float,
        start: tuple,
        end: tuple,
        gamma: float = 0.45,
    ) -> str:
        """Map a single log2FC value to an RGB colour string.

        ``fc_min`` maps to ``start``; ``fc_max`` maps to ``end``.
        Values outside [fc_min, fc_max] are clamped to the extremes.
        ``gamma`` < 1 stretches the middle of the range for better
        perceptual contrast when differences are small.
        """
        t = (fc - fc_min) / (fc_max - fc_min) if fc_max != fc_min else 0.5
        t = max(0.0, min(1.0, t))
        t = t ** gamma
        return "rgb({},{},{})".format(
            int(start[0] + t * (end[0] - start[0])),
            int(start[1] + t * (end[1] - start[1])),
            int(start[2] + t * (end[2] - start[2])),
        )

    # ------------------------------------------------------------------
    # Figure
    # ------------------------------------------------------------------

    def bar_plot(self) -> go.Figure:
        """Builds and displays the two-panel keyword-distribution figure.

        The top panel shows the total count per keyword; the bottom panel
        shows a 100 % stacked bar chart split by compound type.

        Returns:
            The fully configured :class:`plotly.graph_objects.Figure` object.
        """
        by_type = self.keyword_counts_by_type()

        all_keys = sorted(
            set(by_type["Normal"][0])
            | set(by_type["Inhibitor"][0])
            | set(by_type["Inducer"][0]),
            key=lambda v: v.lower(),
        )

        m_norm = self.to_map(by_type["Normal"][0], by_type["Normal"][1])
        m_inh  = self.to_map(by_type["Inhibitor"][0], by_type["Inhibitor"][1])
        m_ind  = self.to_map(by_type["Inducer"][0], by_type["Inducer"][1])

        y_normal_counts  = [m_norm.get(k, 0) for k in all_keys]
        y_inhib_counts   = [m_inh.get(k, 0)  for k in all_keys]
        y_inducer_counts = [m_ind.get(k, 0)  for k in all_keys]

        totals_counts = [
            a + b + c
            for a, b, c in zip(y_normal_counts, y_inhib_counts, y_inducer_counts)
        ]

        y_normal  = self.to_pct(y_normal_counts, totals_counts)
        y_inhib   = self.to_pct(y_inhib_counts,  totals_counts)
        y_inducer = self.to_pct(y_inducer_counts, totals_counts)

        fig = make_subplots(
            rows=2,
            cols=1,
            shared_xaxes=True,
            vertical_spacing=self.vertical_spacing,
            row_heights=self.row_heights,
            subplot_titles=("Total count per keyword", "Keyword distribution"),
        )

        # Top panel – total counts
        fig.add_trace(
            go.Bar(
                x=all_keys,
                y=totals_counts,
                name="Total count",
                marker_color="rgba(80,80,80,0.6)",
                marker_line=dict(color="rgba(80,80,80,1)", width=1),
                text=[str(t) if t > 0 else "" for t in totals_counts],
                textposition="outside",
            ),
            row=1,
            col=1,
        )

        # Bottom panel – 100 % stacked bars
        fig.add_trace(
            go.Bar(x=all_keys, y=y_normal,  name="Normal",
                   marker_color=self.colors["Normal"]),
            row=2, col=1,
        )
        fig.add_trace(
            go.Bar(x=all_keys, y=y_inhib,   name="Inhibitor",
                   marker_color=self.colors["Inhibitor"]),
            row=2, col=1,
        )
        fig.add_trace(
            go.Bar(x=all_keys, y=y_inducer, name="Inducer",
                   marker_color=self.colors["Inducer"]),
            row=2, col=1,
        )

        fig.update_layout(
            barmode="stack",
            height=self.height,
            width=self.width,
            title_text="",
            margin=dict(**self.margin),
            bargap=self.bargap,
        )

        fig.update_xaxes(tickangle=45, automargin=True, row=2, col=1)
        fig.update_yaxes(title_text="Count", automargin=True, row=1, col=1)
        fig.update_yaxes(
            title_text="Percent",
            range=[0, 100],
            ticksuffix="%",
            automargin=True,
            tickmode="array",
            tickvals=[0, 20, 40, 60, 80, 100],
            row=2,
            col=1,
        )
        fig.update_traces(cliponaxis=False, row=1, col=1)

        print(f"df rows: {self.df.shape[0]}")
        fig.show()
        return fig

    def scored_bar_plot(self) -> go.Figure:
        """Plots a scored, gradient-coloured 100 % stacked bar chart per keyword.

        Identical two-panel layout and y-axis (0–100 %) to :meth:`plot`, but
        the bottom panel replaces the plain Inhibitor and Inducer segments with
        per-compound slices. Each compound occupies an equal share of the bar
        (``1 / total_compounds_for_keyword × 100 %``), and its colour encodes
        its log\ :sub:`2` fold-change (log\ :sub:`2`\\ FC) rank:

        * **Inhibitors** – sorted ascending (most inhibitory first, i.e. most
          negative log\ :sub:`2`\\ FC); darkest green at the bottom, lightest
          green at the top.
        * **Inducers** – sorted ascending (weakest induction first); darkest
          red at the bottom, lightest red at the top.
        * **Normal** – single black segment kept intact at the base.

        Two colorbars on the right-hand side map colour shade to the actual
        log\ :sub:`2`\\ FC value for each type.

        Log\ :sub:`2`\\ FC is calculated as
        ``log2(EV_uptake / avg_control_uptake)`` where the average control
        value is library-specific:

        * Prestwick: 0.007618627107104729
        * LOPAC: 0.038644386186536726

        Compounds that do not belong to either library or that have a
        non-positive uptake value are skipped and do not contribute a slice.

        Returns:
            The fully configured :class:`plotly.graph_objects.Figure` object.
        """
        avg_controls = {
            "Prestwick": 0.007618627107104729,
            "LOPAC": 0.038644386186536726,
        }
        uptake_col = "Screen: EV-uptake_Normalized_by_mean"
        library_col = "Library"

        # ── keyword universe ──────────────────────────────────────────────────
        by_type = self.keyword_counts_by_type()
        all_keys = sorted(
            set(by_type["Normal"][0])
            | set(by_type["Inhibitor"][0])
            | set(by_type["Inducer"][0]),
            key=lambda v: v.lower(),
        )

        name_col = "Compound Name"

        # ── per-keyword data ──────────────────────────────────────────────────
        normal_counts: dict = {}
        inh_data: dict = {}   # sorted list of (log2FC, name, ev_uptake) per keyword
        ind_data: dict = {}
        total_counts: dict = {}

        # For each keyword, extract the subset of compounds and compute the 
        # log2FC values for Inhibitors and Inducers.
        for kw in all_keys:
            df_kw = Graph._compounds_for_keyword(self.df, self.keyword_col, kw)
            type_s = df_kw[self.type_col].astype(str).str.lower()
            n_norm = int((type_s == "normal").sum())
            i_data = Graph._compound_data_list(
                df_kw[type_s == "inhibitor"], avg_controls, uptake_col, library_col, name_col
            )
            d_data = Graph._compound_data_list(
                df_kw[type_s == "inducer"], avg_controls, uptake_col, library_col, name_col
            )
            normal_counts[kw] = n_norm
            inh_data[kw] = i_data
            ind_data[kw] = d_data
            total_counts[kw] = n_norm + len(i_data) + len(d_data)

        max_inh = max((len(v) for v in inh_data.values()), default=0)
        max_ind = max((len(v) for v in ind_data.values()), default=0)

        # Global log2FC values – colours are mapped against the 5th–95th
        # percentile range so the full gradient spans the bulk of the data
        # rather than being compressed by extreme outliers.
        all_inh_fc = [t[0] for data in inh_data.values() for t in data]
        all_ind_fc = [t[0] for data in ind_data.values() for t in data]
        inh_min = min(all_inh_fc) if all_inh_fc else -1.0
        inh_max = max(all_inh_fc) if all_inh_fc else 0.0
        ind_min = min(all_ind_fc) if all_ind_fc else 0.0
        ind_max = max(all_ind_fc) if all_ind_fc else 1.0

        # Percentile-based scale endpoints (colour mapping + colorbars)
        _p = 5
        inh_scale_min = Graph._percentile_val(all_inh_fc, _p)     if len(all_inh_fc) >= 4 else inh_min
        inh_scale_max = Graph._percentile_val(all_inh_fc, 100-_p) if len(all_inh_fc) >= 4 else inh_max
        ind_scale_min = Graph._percentile_val(all_ind_fc, _p)     if len(all_ind_fc) >= 4 else ind_min
        ind_scale_max = Graph._percentile_val(all_ind_fc, 100-_p) if len(all_ind_fc) >= 4 else ind_max

        # ── figure ────────────────────────────────────────────────────────────
        all_totals = [total_counts[k] for k in all_keys]

        fig = make_subplots(
            rows=2,
            cols=1,
            shared_xaxes=True,
            vertical_spacing=self.vertical_spacing,
            row_heights=self.row_heights,
            subplot_titles=("Total count per keyword", "Scored keyword distribution"),
        )

        # Top panel – total compound count per keyword
        fig.add_trace(
            go.Bar(
                x=all_keys,
                y=all_totals,
                name="Total count",
                marker_color="rgba(80,80,80,0.6)",
                marker_line=dict(color="rgba(80,80,80,1)", width=1),
                text=[str(t) if t > 0 else "" for t in all_totals],
                textposition="outside",
                showlegend=False,
            ),
            row=1,
            col=1,
        )

        # Bottom panel – Normal base (intact, single black segment per keyword)
        fig.add_trace(
            go.Bar(
                x=all_keys,
                y=[
                    normal_counts[k] / total_counts[k] * 100
                    if total_counts[k] else 0
                    for k in all_keys
                ],
                name="Normal",
                marker_color="black",
                showlegend=True,
                hovertemplate="<b>%{x}</b><br>Normal compounds: %{y:.1f}%<extra></extra>",
            ),
            row=2,
            col=1,
        )

        # Bottom panel – Inhibitor gradient layers.
        # Color per compound: most negative log2FC → lightest green (t=0),
        # least negative → darkest green (t=1), mapped against global inh range.
        for i in range(max_inh):
            y_vals = []
            colors = []
            customdata = []
            for k in all_keys:
                if i < len(inh_data[k]) and total_counts[k]:
                    fc, cname, ev = inh_data[k][i]
                    y_vals.append(1 / total_counts[k] * 100)
                    colors.append(Graph._interpolate_color(fc, inh_scale_min, inh_scale_max,
                                                           (210, 255, 210), (0, 60, 0)))
                    customdata.append([cname, ev, fc])
                else:
                    y_vals.append(0)
                    colors.append("rgba(0,0,0,0)")
                    customdata.append(["", 0.0, 0.0])
            fig.add_trace(
                go.Bar(
                    x=all_keys,
                    y=y_vals,
                    name="Inhibitor" if i == 0 else "",
                    legendgroup="inhibitor",
                    showlegend=(i == 0),
                    marker_color=colors,
                    customdata=customdata,
                    hovertemplate=(
                        "<b>%{x}</b><br>"
                        "Compound: %{customdata[0]}<br>"
                        "EV uptake: %{customdata[1]:.6f}<br>"
                        "log₂FC: %{customdata[2]:.3f}"
                        "<extra></extra>"
                    ),
                ),
                row=2,
                col=1,
            )

        # Bottom panel – Inducer gradient layers.
        # Color per compound: least positive log2FC → darkest red (t=0),
        # most positive → lightest red (t=1), mapped against global ind range.
        for i in range(max_ind):
            y_vals = []
            colors = []
            customdata = []
            for k in all_keys:
                if i < len(ind_data[k]) and total_counts[k]:
                    fc, cname, ev = ind_data[k][i]
                    y_vals.append(1 / total_counts[k] * 100)
                    colors.append(Graph._interpolate_color(fc, ind_scale_min, ind_scale_max,
                                                           (180, 0, 0), (255, 210, 210)))
                    customdata.append([cname, ev, fc])
                else:
                    y_vals.append(0)
                    colors.append("rgba(0,0,0,0)")
                    customdata.append(["", 0.0, 0.0])
            fig.add_trace(
                go.Bar(
                    x=all_keys,
                    y=y_vals,
                    name="Inducer" if i == 0 else "",
                    legendgroup="inducer",
                    showlegend=(i == 0),
                    marker_color=colors,
                    customdata=customdata,
                    hovertemplate=(
                        "<b>%{x}</b><br>"
                        "Compound: %{customdata[0]}<br>"
                        "EV uptake: %{customdata[1]:.6f}<br>"
                        "log₂FC: %{customdata[2]:.3f}"
                        "<extra></extra>"
                    ),
                ),
                row=2,
                col=1,
            )

        # Colorbars – dummy invisible scatter traces that carry the scale legend.
        # Tick values: floor(min), midpoint, ceil(max) – guarantees full range coverage.
        inh_tick_min = math.floor(inh_min * 10) / 10
        inh_tick_max = math.ceil(inh_max * 10) / 10
        inh_tick_mid = round((inh_tick_min + inh_tick_max) / 2, 1)
        ind_tick_min = math.floor(ind_min * 10) / 10
        ind_tick_max = math.ceil(ind_max * 10) / 10
        ind_tick_mid = round((ind_tick_min + ind_tick_max) / 2, 1)

        # Inhibitor: light green (most negative / strongest) → dark green (weakest).
        fig.add_trace(
            go.Scatter(
                x=[None],
                y=[None],
                mode="markers",
                marker=dict(
                    colorscale=[[0, "rgb(210,255,210)"], [1, "rgb(0,60,0)"]],
                    cmin=inh_tick_min,
                    cmax=inh_tick_max,
                    color=[inh_tick_min],
                    showscale=True,
                    colorbar=dict(
                        title=dict(text="Inhibitor<br>log₂FC", side="right"),
                        x=1.02,
                        y=0.25,
                        len=0.45,
                        thickness=14,
                        tickmode="array",
                        tickvals=[inh_tick_min, inh_tick_mid, inh_tick_max],
                        tickformat=".1f",
                        outlinewidth=1,
                    ),
                ),
                showlegend=False,
                hoverinfo="skip",
            ),
            row=1,
            col=1,
        )

        # Inducer: dark red (weakest / lowest log2FC) → light red (strongest).
        fig.add_trace(
            go.Scatter(
                x=[None],
                y=[None],
                mode="markers",
                marker=dict(
                    colorscale=[[0, "rgb(180,0,0)"], [1, "rgb(255,210,210)"]],
                    cmin=ind_tick_min,
                    cmax=ind_tick_max,
                    color=[ind_tick_min],
                    showscale=True,
                    colorbar=dict(
                        title=dict(text="Inducer<br>log₂FC", side="right"),
                        x=1.13,
                        y=0.25,
                        len=0.45,
                        thickness=14,
                        tickmode="array",
                        tickvals=[ind_tick_min, ind_tick_mid, ind_tick_max],
                        tickformat=".1f",
                        outlinewidth=1,
                    ),
                ),
                showlegend=False,
                hoverinfo="skip",
            ),
            row=1,
            col=1,
        )

        fig.update_layout(
            barmode="stack",
            height=self.height,
            width=self.width,
            title_text="",
            margin=dict(**self.margin),
            bargap=self.bargap,
        )

        fig.update_xaxes(tickangle=45, automargin=True, row=2, col=1)
        fig.update_yaxes(title_text="Count", automargin=True, row=1, col=1)
        fig.update_yaxes(
            title_text="Percent",
            range=[0, 100],
            ticksuffix="%",
            automargin=True,
            tickmode="array",
            tickvals=[0, 20, 40, 60, 80, 100],
            row=2,
            col=1,
        )
        fig.update_traces(cliponaxis=False, row=1, col=1)

        print(f"df rows: {self.df.shape[0]}")
        fig.show()
        return fig
    

    def scored_table(self) -> pd.DataFrame:
        """Returns a DataFrame matching the compound order in :meth:`scored_bar_plot`.

        Each column is a keyword (Group). Each row contains the compound name
        for that position in the stacked bar, ordered top-to-bottom as in the
        chart:

        1. Strongest inducer first (highest log₂FC).
        2. Progressively weaker inducers.
        3. Weakest inhibitor (least negative log₂FC, closest to 0).
        4. Progressively stronger inhibitors down to the most negative log₂FC.

        Only Inhibitor and Inducer compounds are included; Normal (control)
        compounds are excluded. Keywords with fewer compounds than the longest
        column are padded with ``None``.

        Returns:
            A :class:`pandas.DataFrame` whose columns are keyword names and
            whose rows are compound names in the described order.
        """
        avg_controls = {
            "Prestwick": 0.007618627107104729,
            "LOPAC": 0.038644386186536726,
        }
        uptake_col = "Screen: EV-uptake_Normalized_by_mean"
        library_col = "Library"
        name_col = "Compound Name"

        by_type = self.keyword_counts_by_type()
        all_keys = sorted(
            set(by_type["Normal"][0])
            | set(by_type["Inhibitor"][0])
            | set(by_type["Inducer"][0]),
            key=lambda v: v.lower(),
        )

        columns: dict[str, list] = {}
        for kw in all_keys:
            df_kw = Graph._compounds_for_keyword(self.df, self.keyword_col, kw)
            type_s = df_kw[self.type_col].astype(str).str.lower()

            # sorted ascending: inh_list[0] = most negative (strongest inhibitor)
            inh_list = Graph._compound_data_list(
                df_kw[type_s == "inhibitor"], avg_controls, uptake_col, library_col, name_col
            )
            # sorted ascending: ind_list[0] = weakest inducer, ind_list[-1] = strongest
            ind_list = Graph._compound_data_list(
                df_kw[type_s == "inducer"], avg_controls, uptake_col, library_col, name_col
            )

            # Top-to-bottom order in the chart:
            #   strongest inducer → weakest inducer → weakest inhibitor → strongest inhibitor
            ordered_names = (
                [item[1] for item in reversed(ind_list)]   # inducers: strongest first
                + [item[1] for item in reversed(inh_list)] # inhibitors: weakest first, most negative last
            )
            columns[kw] = ordered_names

        max_len = max((len(v) for v in columns.values()), default=0)
        for kw in all_keys:
            lst = columns[kw]
            lst += [None] * (max_len - len(lst))

        return pd.DataFrame(columns)

    def scored_color_table(self) -> pd.DataFrame:
        """Returns a DataFrame of RGB colour strings matching :meth:`scored_table`.

        Same shape and column/row order as :meth:`scored_table`. Each cell
        contains the ``"rgb(r,g,b)"`` string that would be used to colour the
        corresponding compound slice in :meth:`scored_bar_plot`, or ``None``
        for padding positions. Use :meth:`to_excel` to apply these colours as
        cell background fills in an Excel workbook.

        Returns:
            A :class:`pandas.DataFrame` of ``"rgb(r,g,b)"`` strings (or
            ``None``) with the same shape as :meth:`scored_table`.
        """
        avg_controls = {
            "Prestwick": 0.007618627107104729,
            "LOPAC": 0.038644386186536726,
        }
        uptake_col = "Screen: EV-uptake_Normalized_by_mean"
        library_col = "Library"
        name_col = "Compound Name"

        by_type = self.keyword_counts_by_type()
        all_keys = sorted(
            set(by_type["Normal"][0])
            | set(by_type["Inhibitor"][0])
            | set(by_type["Inducer"][0]),
            key=lambda v: v.lower(),
        )

        # Collect per-keyword data and global FC lists for the percentile scale
        kw_data: dict = {}
        all_inh_fc: list = []
        all_ind_fc: list = []
        for kw in all_keys:
            df_kw = Graph._compounds_for_keyword(self.df, self.keyword_col, kw)
            type_s = df_kw[self.type_col].astype(str).str.lower()
            inh_list = Graph._compound_data_list(
                df_kw[type_s == "inhibitor"], avg_controls, uptake_col, library_col, name_col
            )
            ind_list = Graph._compound_data_list(
                df_kw[type_s == "inducer"], avg_controls, uptake_col, library_col, name_col
            )
            kw_data[kw] = (inh_list, ind_list)
            all_inh_fc.extend(t[0] for t in inh_list)
            all_ind_fc.extend(t[0] for t in ind_list)

        _p = 5
        inh_min = min(all_inh_fc) if all_inh_fc else -1.0
        inh_max = max(all_inh_fc) if all_inh_fc else 0.0
        ind_min = min(all_ind_fc) if all_ind_fc else 0.0
        ind_max = max(all_ind_fc) if all_ind_fc else 1.0
        inh_scale_min = Graph._percentile_val(all_inh_fc, _p)     if len(all_inh_fc) >= 4 else inh_min
        inh_scale_max = Graph._percentile_val(all_inh_fc, 100-_p) if len(all_inh_fc) >= 4 else inh_max
        ind_scale_min = Graph._percentile_val(all_ind_fc, _p)     if len(all_ind_fc) >= 4 else ind_min
        ind_scale_max = Graph._percentile_val(all_ind_fc, 100-_p) if len(all_ind_fc) >= 4 else ind_max

        columns: dict[str, list] = {}
        for kw in all_keys:
            inh_list, ind_list = kw_data[kw]
            ordered_colors = (
                [
                    Graph._interpolate_color(
                        fc, ind_scale_min, ind_scale_max, (180, 0, 0), (255, 210, 210)
                    )
                    for fc, _, _ in reversed(ind_list)
                ]
                + [
                    Graph._interpolate_color(
                        fc, inh_scale_min, inh_scale_max, (210, 255, 210), (0, 60, 0)
                    )
                    for fc, _, _ in reversed(inh_list)
                ]
            )
            columns[kw] = ordered_colors

        max_len = max((len(v) for v in columns.values()), default=0)
        for kw in all_keys:
            columns[kw] += [None] * (max_len - len(columns[kw]))

        return pd.DataFrame(columns)

    def to_excel(self, path: str, table_df: pd.DataFrame | None = None) -> None:
        """Write the scored compound table to an Excel file with bar-matched cell colours.

        Each cell's background fill is the exact RGB colour used for that
        compound slice in :meth:`scored_bar_plot`, computed via
        :meth:`scored_color_table`. Text on dark backgrounds is rendered white
        for readability.

        Args:
            path: Destination ``.xlsx`` file path.
            table_df: Compound name table to write. Accepts the output of
                :meth:`scored_table`, :meth:`compound_finder`, or
                :meth:`compound_ranking`. Defaults to :meth:`scored_table`.

        Raises:
            ImportError: If ``openpyxl`` is not installed.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for to_excel(). Install it with: pip install openpyxl"
            ) from exc

        if table_df is None:
            table_df = self.scored_table()
        color_df = self.scored_color_table()

        def _rgb_to_argb(rgb_str: str) -> str:
            """Convert ``"rgb(r,g,b)"`` to openpyxl ARGB hex ``"FFrrggbb"``."""
            m = re.match(r"rgb\((\d+),(\d+),(\d+)\)", rgb_str.replace(" ", ""))
            if not m:
                return "FFFFFFFF"
            r, g, b = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return f"FF{r:02X}{g:02X}{b:02X}"

        wb = Workbook()
        ws = wb.active

        # Header row
        for col_idx, col_name in enumerate(table_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)

        # Data rows
        n_color_rows = len(color_df)
        for row_idx, row_vals in enumerate(table_df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                color_row = row_idx - 2
                color_col = col_idx - 1
                if color_row < n_color_rows:
                    rgb_str = color_df.iloc[color_row, color_col]
                else:
                    rgb_str = None
                if isinstance(rgb_str, str):
                    cell.fill = PatternFill(fill_type="solid", fgColor=_rgb_to_argb(rgb_str))
                    m = re.match(r"rgb\((\d+),(\d+),(\d+)\)", rgb_str.replace(" ", ""))
                    if m:
                        luminance = 0.299 * int(m.group(1)) + 0.587 * int(m.group(2)) + 0.114 * int(m.group(3))
                        if luminance < 128:
                            cell.font = Font(color="FFFFFFFF")

        wb.save(path)
        print(f"Excel saved: {path}")

    def mark_excel(
        self,
        input_path: str,
        output_path: str,
        compounds: list[str],
    ) -> None:
        """Mark matching compounds in an existing coloured Excel file with ``[X]``.

        Loads *input_path* (typically produced by :meth:`to_excel`) and
        appends ``" [X]"`` to every cell whose text value contains a name from
        *compounds*. All cell background colours, fonts, borders, and column
        widths are preserved exactly — nothing is touched except the text of
        matching cells.

        Args:
            input_path:  Path to an existing ``.xlsx`` file (e.g. the output
                of :meth:`to_excel`).
            output_path: Destination path for the annotated file. May equal
                *input_path* to overwrite in place.
            compounds:   List of compound name strings to search for. Matching
                is substring-based (same logic as :meth:`compound_finder`).

        Raises:
            ImportError: If ``openpyxl`` is not installed.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for mark_excel(). Install with: pip install openpyxl"
            ) from exc

        compound_set = set(compounds)
        wb = load_workbook(input_path)
        ws = wb.active

        n_matches = 0
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and any(term in val for term in compound_set):
                    cell.value = val + " [X]"
                    n_matches += 1

        wb.save(output_path)
        print(f"Matches marked: {n_matches}")
        print(f"Saved: {output_path}")

    def compound_finder(
        self,
        compounds: list[str],
        input_path: str | None = None,
        output_path: str | None = None,
    ) -> pd.DataFrame:
        """Search for compounds in the scored table and mark matches with ``[X]``.

        Calls :meth:`scored_table` to obtain the ranked compound table, then
        scans every cell. Any cell whose value contains a name from *compounds*
        is suffixed with ``" [X]"``. All other cells are left unchanged.
        ``None`` padding cells are never modified.

        When *input_path* is supplied the method additionally loads that
        ``.xlsx`` file (typically produced by :meth:`to_excel`) and appends
        ``" [X]"`` to matching cells **without touching any cell formatting**
        (colours, fonts, borders). The annotated workbook is saved to
        *output_path*, or back to *input_path* if *output_path* is omitted.
        This means the coloured Excel and the returned DataFrame are always
        consistent.

        Args:
            compounds: List of compound name strings to search for. Matching
                is substring-based (same as :meth:`mark_excel`).
            input_path: Optional path to an existing coloured ``.xlsx`` file.
                When provided the file is marked in-place and saved.
            output_path: Destination path for the annotated Excel file.
                Defaults to *input_path* (overwrites the source file).

        Returns:
            A copy of the :meth:`scored_table` DataFrame with matching compound
            names suffixed by ``" [X]"``.

        Raises:
            ImportError: If *input_path* is provided and ``openpyxl`` is not
                installed.
        """
        df = self.scored_table()
        compound_set = set(compounds)

        def mark(val):
            if val is not None and any(term in val for term in compound_set):
                return val + " [X]"
            return val

        marked = df.apply(lambda col: col.map(mark))
        n_matches = marked.apply(
            lambda col: col.map(lambda v: v is not None and v.endswith(" [X]"))
        ).values.sum()
        print(f"Matches found: {n_matches}")

        if input_path is not None:
            self.mark_excel(input_path, output_path or input_path, compounds)

        return marked

    

    def compound_ranking(self, marked: pd.DataFrame) -> pd.DataFrame:
        """Rank inhibitors and inducers by EV-uptake and annotate the marked table.

        Builds two temporary ranked DataFrames from the master data:

        * **Inhibitors** – sorted lowest-to-highest uptake; rank 1 is the
          strongest inhibitor (lowest uptake value).
        * **Inducers** – sorted highest-to-lowest uptake; rank 1 is the
          strongest inducer (highest uptake value).

        Both DataFrames are merged into a single lookup table. Every cell in
        the *marked* table that carries a ``[X]`` suffix has its ``[X]``
        stripped to recover the bare compound name, which is then looked up in
        the ranking table. When a match is found the cell receives an
        additional label in the format ``[Library, Function, Rank]`` appended
        after the existing ``[X]``.

        Library abbreviations: ``LOP`` (LOPAC), ``Prest`` (Prestwick).
        Function abbreviations: ``Inh`` (Inhibitor), ``Ind`` (Inducer).

        Example cell value after labelling::

            "Cytochalasin D [X] [LOP, Ind, 2]"

        Args:
            marked: DataFrame returned by :meth:`compound_finder`, whose cells
                may carry ``" [X]"`` suffixes.

        Returns:
            A copy of *marked* with matching cells additionally suffixed by
            the ranking label. Cells without ``[X]`` are left untouched.
        """
        uptake_col = "Screen: EV-uptake_Normalized_by_mean"
        library_col = "Library"
        name_col = "Compound Name"
        type_col = "Screen: Effect on EV uptake"

        lib_abbr_map = {"LOPAC": "LOP", "Prestwick": "Prest"}
        func_abbr_map = {"Inhibitor": "Inh", "Inducer": "Ind"}

        # ── Inhibitors: rank 1 = lowest uptake (strongest inhibitor) ──────────
        df_inh = (
            self.df[self.df[type_col].astype(str).str.lower() == "inhibitor"]
            [[name_col, uptake_col, library_col, type_col]]
            .copy()
            .sort_values(uptake_col, ascending=True)
            .reset_index(drop=True)
        )
        df_inh["Rank"] = range(1, len(df_inh) + 1)

        # ── Inducers: rank 1 = highest uptake (strongest inducer) ─────────────
        df_ind = (
            self.df[self.df[type_col].astype(str).str.lower() == "inducer"]
            [[name_col, uptake_col, library_col, type_col]]
            .copy()
            .sort_values(uptake_col, ascending=False)
            .reset_index(drop=True)
        )
        df_ind["Rank"] = range(1, len(df_ind) + 1)

        # ── Merge and build name → label lookup ───────────────────────────────
        df_ranked = pd.concat([df_inh, df_ind], ignore_index=True)

        ranking_lookup: dict[str, str] = {}
        for _, row in df_ranked.iterrows():
            cname = str(row[name_col])
            lib = lib_abbr_map.get(str(row[library_col]), str(row[library_col]))
            func = func_abbr_map.get(str(row[type_col]), str(row[type_col]))
            rank = int(row["Rank"])
            ranking_lookup[cname] = f"[{lib}, {func}, {rank}]"

        # ── Annotate the marked table ──────────────────────────────────────────
        def add_ranking_label(val):
            if not isinstance(val, str):
                return val
            # Strip [X] (if present) to recover the bare compound name
            bare_name = val.replace(" [X]", "").strip()
            rank_label = ranking_lookup.get(bare_name)
            if rank_label is not None:
                return val + " " + rank_label
            return val

        result = marked.copy()
        result = result.apply(lambda col: col.map(add_ranking_label))
        n_labelled = result.apply(
            lambda col: col.map(
                lambda v: isinstance(v, str) and re.search(r"\[.+,.+,\s*\d+\]", v) is not None
            )
        ).values.sum()
        print(f"Ranking labels added: {n_labelled}")
        return result

        


    def outliers_plot(
        self,
        df_outliers: pd.DataFrame,
        df_normal: pd.DataFrame,
        *,
        outlier_label: str = "df_outliers",
        normal_label: str = "df_normal_comp",
        outlier_color: str = "steelblue",
        normal_color: str = "indianred",
        title: str = "Keyword Frequency (alphabetical x-axis)",
        vertical_spacing: float = 0.3,
    ) -> go.Figure:
        """Plots keyword-count bar charts for outlier and normal compounds side-by-side.

        Creates a two-row subplot: the top panel shows keyword counts for
        ``df_outliers`` and the bottom panel shows counts for ``df_normal``.
        Each bar is annotated with the keyword's percentage of that DataFrame's
        total row count.

        Args:
            df_outliers: DataFrame containing outlier (activator/inhibitor) compounds.
            df_normal: DataFrame containing normal (inlier) compounds.
            outlier_label: Legend / subplot title label for the outlier series.
            normal_label: Legend / subplot title label for the normal series.
            outlier_color: Bar colour for the outlier panel.
            normal_color: Bar colour for the normal panel.
            title: Overall figure title.
            vertical_spacing: Vertical gap between the two subplot rows.

        Returns:
            The fully configured :class:`plotly.graph_objects.Figure` object.
        """
        keys_out, counts_out, perc_out = self.keyword_counts(df_outliers)
        keys_norm, counts_norm, perc_norm = self.keyword_counts(df_normal)

        fig = make_subplots(
            rows=2,
            cols=1,
            shared_xaxes=False,
            vertical_spacing=vertical_spacing,
            subplot_titles=(
                f"{outlier_label}: Keyword counts",
                f"{normal_label}: Keyword counts",
            ),
        )

        fig.add_trace(
            go.Bar(
                x=keys_out,
                y=counts_out,
                marker_color=outlier_color,
                name=outlier_label,
                text=[f"{p:.2f}%" for p in perc_out],
                textposition="outside",
            ),
            row=1,
            col=1,
        )

        fig.add_trace(
            go.Bar(
                x=keys_norm,
                y=counts_norm,
                marker_color=normal_color,
                name=normal_label,
                text=[f"{p:.2f}%" for p in perc_norm],
                textposition="outside",
            ),
            row=2,
            col=1,
        )

        fig.update_layout(
            height=self.height,
            width=self.width,
            showlegend=False,
            title_text=title,
            margin=dict(**self.margin),
            bargap=self.bargap,
        )

        fig.update_xaxes(tickangle=45, automargin=True)
        fig.update_yaxes(title_text="Count", automargin=True)

        print(f"{outlier_label} rows: {df_outliers.shape[0]}")
        print(f"{normal_label} rows: {df_normal.shape[0]}")
        fig.show()
        return fig
